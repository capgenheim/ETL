"""
Unit tests for the Transformation app.
Tests header extraction service and API endpoints.
"""
import csv
import io
import os
import tempfile

from django.test import TestCase, override_settings
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APIClient

import openpyxl

from apps.accounts.models import User
from apps.transformation.models import UploadedFile
from apps.transformation.services import (
    extract_headers,
    HeaderExtractionError,
    get_file_extension,
)

TEMP_MEDIA = tempfile.mkdtemp()


# ─── Header Extraction Tests ───────────────────────────────────────────

class TestGetFileExtension(TestCase):
    def test_xlsx(self):
        self.assertEqual(get_file_extension('report.xlsx'), 'xlsx')

    def test_xls(self):
        self.assertEqual(get_file_extension('legacy.XLS'), 'xls')

    def test_csv(self):
        self.assertEqual(get_file_extension('data.csv'), 'csv')

    def test_unsupported(self):
        self.assertEqual(get_file_extension('image.png'), 'png')


class TestExtractHeadersCsv(TestCase):
    def test_extracts_csv_headers(self):
        content = 'Name,Age,Email,Department\nJohn,30,john@test.com,IT'
        f = io.BytesIO(content.encode('utf-8'))
        result = extract_headers(f, 'test.csv')
        self.assertEqual(result['headers'], ['Name', 'Age', 'Email', 'Department'])
        self.assertEqual(result['field_count'], 4)
        self.assertEqual(result['file_format'], 'csv')

    def test_csv_with_bom(self):
        # Realistic BOM-prefixed CSV (as Windows Excel would create)
        content = 'ID,Name,Value'
        f = io.BytesIO(content.encode('utf-8-sig'))
        result = extract_headers(f, 'bom.csv')
        self.assertEqual(result['headers'], ['ID', 'Name', 'Value'])

    def test_csv_empty(self):
        f = io.BytesIO(b'')
        result = extract_headers(f, 'empty.csv')
        self.assertEqual(result['headers'], [])
        self.assertEqual(result['field_count'], 0)

    def test_csv_blanks_filtered(self):
        """Blank headers in the middle or end should be excluded."""
        content = 'Name,,Age,,Email,'
        f = io.BytesIO(content.encode('utf-8'))
        result = extract_headers(f, 'gaps.csv')
        self.assertEqual(result['headers'], ['Name', 'Age', 'Email'])
        self.assertEqual(result['field_count'], 3)


class TestExtractHeadersXlsx(TestCase):
    def _make_xlsx(self, headers):
        """Create an in-memory xlsx file with given headers."""
        wb = openpyxl.Workbook()
        ws = wb.active
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        # Add a data row to ensure only headers are extracted
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col, value=f'data_{col}')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_extracts_xlsx_headers(self):
        headers = ['Account No', 'Balance', 'Currency', 'Date']
        f = self._make_xlsx(headers)
        result = extract_headers(f, 'report.xlsx')
        self.assertEqual(result['headers'], headers)
        self.assertEqual(result['field_count'], 4)
        self.assertEqual(result['file_format'], 'xlsx')

    def test_xlsx_empty_workbook(self):
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = extract_headers(buf, 'empty.xlsx')
        self.assertEqual(result['headers'], [])


class TestExtractHeadersValidation(TestCase):
    def test_unsupported_format_raises(self):
        f = io.BytesIO(b'not-an-image')
        with self.assertRaises(HeaderExtractionError) as ctx:
            extract_headers(f, 'file.png')
        self.assertIn('Unsupported format', str(ctx.exception))


# ─── API Tests ──────────────────────────────────────────────────────────

@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class TestFileUploadAPI(TestCase):
    def setUp(self):
        from oauth2_provider.models import Application, AccessToken
        from django.utils import timezone
        import datetime

        self.user = User.objects.create_user(
            username='tester',
            email='tester@etl.local',
            password='Test@12345',
        )
        self.app = Application.objects.create(
            name='test-app',
            client_type=Application.CLIENT_CONFIDENTIAL,
            authorization_grant_type=Application.GRANT_PASSWORD,
            user=self.user,
        )
        self.token = AccessToken.objects.create(
            user=self.user,
            application=self.app,
            token='test-token-upload-123',
            expires=timezone.now() + datetime.timedelta(hours=1),
            scope='read write',
        )
        self.client = APIClient()
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {self.token.token}')

    def _make_csv_file(self, name='test.csv', headers=None):
        if headers is None:
            headers = ['Name', 'Age', 'Email']
        content = ','.join(headers) + '\nJohn,30,john@test.com'
        return SimpleUploadedFile(name, content.encode('utf-8'), content_type='text/csv')

    def _make_xlsx_file(self, name='test.xlsx', headers=None):
        if headers is None:
            headers = ['Col_A', 'Col_B', 'Col_C']
        wb = openpyxl.Workbook()
        ws = wb.active
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        buf = io.BytesIO()
        wb.save(buf)
        return SimpleUploadedFile(name, buf.getvalue(),
                                  content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_upload_csv_source(self):
        f = self._make_csv_file()
        resp = self.client.post('/api/transformation/upload/', {
            'files': [f],
            'file_type': 'source',
        }, format='multipart')
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(len(resp.data['uploaded']), 1)
        uploaded = resp.data['uploaded'][0]
        self.assertEqual(uploaded['file_type'], 'source')
        self.assertEqual(uploaded['field_count'], 3)
        self.assertEqual(uploaded['headers_json'], ['Name', 'Age', 'Email'])

    def test_upload_xlsx_canvas(self):
        f = self._make_xlsx_file(headers=['Balance', 'Date', 'Account'])
        resp = self.client.post('/api/transformation/upload/', {
            'files': [f],
            'file_type': 'canvas',
        }, format='multipart')
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.data['uploaded'][0]['file_format'], 'xlsx')
        self.assertEqual(resp.data['uploaded'][0]['headers_json'], ['Balance', 'Date', 'Account'])

    def test_upload_multiple_files(self):
        f1 = self._make_csv_file('file1.csv', ['A', 'B'])
        f2 = self._make_csv_file('file2.csv', ['X', 'Y', 'Z'])
        resp = self.client.post('/api/transformation/upload/', {
            'files': [f1, f2],
            'file_type': 'source',
        }, format='multipart')
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(len(resp.data['uploaded']), 2)

    def test_upload_no_file_type_returns_400(self):
        f = self._make_csv_file()
        resp = self.client.post('/api/transformation/upload/', {
            'files': [f],
        }, format='multipart')
        self.assertEqual(resp.status_code, 400)

    def test_upload_invalid_format(self):
        f = SimpleUploadedFile('bad.png', b'fake-image', content_type='image/png')
        resp = self.client.post('/api/transformation/upload/', {
            'files': [f],
            'file_type': 'source',
        }, format='multipart')
        self.assertEqual(resp.status_code, 400)
        self.assertIn('errors', resp.data)

    def test_upload_duplicate_rejected(self):
        """Re-uploading a file with the same name and type should return a duplicate error."""
        f1 = self._make_csv_file('report.csv', ['A', 'B'])
        resp1 = self.client.post('/api/transformation/upload/', {
            'files': [f1],
            'file_type': 'source',
        }, format='multipart')
        self.assertEqual(resp1.status_code, 201)

        # Upload same filename again for the same type
        f2 = self._make_csv_file('report.csv', ['X', 'Y'])
        resp2 = self.client.post('/api/transformation/upload/', {
            'files': [f2],
            'file_type': 'source',
        }, format='multipart')
        self.assertEqual(resp2.status_code, 400)  # No successful uploads
        self.assertEqual(len(resp2.data.get('errors', [])), 1)
        self.assertIn('Duplicate', resp2.data['errors'][0]['error'])

    def test_same_filename_different_type_allowed(self):
        """Same filename is allowed across different types (source vs canvas)."""
        f1 = self._make_csv_file('data.csv', ['A', 'B'])
        resp1 = self.client.post('/api/transformation/upload/', {
            'files': [f1],
            'file_type': 'source',
        }, format='multipart')
        self.assertEqual(resp1.status_code, 201)

        f2 = self._make_csv_file('data.csv', ['X', 'Y'])
        resp2 = self.client.post('/api/transformation/upload/', {
            'files': [f2],
            'file_type': 'canvas',
        }, format='multipart')
        self.assertEqual(resp2.status_code, 201)  # Different type, should succeed


@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class TestFileListAndDeleteAPI(TestCase):
    def setUp(self):
        from oauth2_provider.models import Application, AccessToken
        from django.utils import timezone
        import datetime

        self.user = User.objects.create_user(
            username='lister', email='lister@etl.local', password='Test@12345',
        )
        self.app = Application.objects.create(
            name='test-app-list', client_type=Application.CLIENT_CONFIDENTIAL,
            authorization_grant_type=Application.GRANT_PASSWORD, user=self.user,
        )
        self.token = AccessToken.objects.create(
            user=self.user, application=self.app,
            token='test-token-list-456',
            expires=timezone.now() + datetime.timedelta(hours=1),
            scope='read write',
        )
        self.client = APIClient()
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {self.token.token}')

        # Seed files
        self.source_file = UploadedFile.objects.create(
            file_type='source', original_filename='data.csv', file_format='csv',
            headers_json=['A', 'B', 'C'], field_count=3,
            uploaded_by=self.user, file='uploads/source/data.csv',
        )
        self.canvas_file = UploadedFile.objects.create(
            file_type='canvas', original_filename='canvas.xlsx', file_format='xlsx',
            headers_json=['X', 'Y'], field_count=2,
            uploaded_by=self.user, file='uploads/canvas/canvas.xlsx',
        )

    def test_list_all_files(self):
        resp = self.client.get('/api/transformation/files/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.data), 2)

    def test_list_source_files_only(self):
        resp = self.client.get('/api/transformation/files/?type=source')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.data), 1)
        self.assertEqual(resp.data[0]['file_type'], 'source')

    def test_list_canvas_files_only(self):
        resp = self.client.get('/api/transformation/files/?type=canvas')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.data), 1)
        self.assertEqual(resp.data[0]['file_type'], 'canvas')

    def test_delete_file(self):
        resp = self.client.delete(f'/api/transformation/files/{self.source_file.id}/')
        self.assertEqual(resp.status_code, 204)
        self.assertFalse(UploadedFile.objects.filter(id=self.source_file.id).exists())


# ─── Package CRUD Tests ────────────────────────────────────────────────

@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class TestPackageCRUDAPI(TestCase):
    def setUp(self):
        from oauth2_provider.models import Application, AccessToken
        from django.utils import timezone
        import datetime

        self.user = User.objects.create_user(
            username='pkguser', email='pkguser@etl.local', password='Test@12345',
            first_name='John', last_name='Doe',
        )
        self.app = Application.objects.create(
            name='test-app-pkg', client_type=Application.CLIENT_CONFIDENTIAL,
            authorization_grant_type=Application.GRANT_PASSWORD, user=self.user,
        )
        self.token = AccessToken.objects.create(
            user=self.user, application=self.app,
            token='test-token-pkg-789',
            expires=timezone.now() + datetime.timedelta(hours=1),
            scope='read write',
        )
        self.client = APIClient()
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {self.token.token}')

        # Seed source and canvas files
        self.source = UploadedFile.objects.create(
            file_type='source', original_filename='bank.csv', file_format='csv',
            headers_json=['Date', 'Description', 'Amount', 'Balance'],
            field_count=4, uploaded_by=self.user,
        )
        self.canvas = UploadedFile.objects.create(
            file_type='canvas', original_filename='output.xlsx', file_format='xlsx',
            headers_json=['TransDate', 'Desc', 'Value'],
            field_count=3, uploaded_by=self.user,
        )

    def _create_package(self, **overrides):
        """Helper to create a package via API."""
        data = {
            'name': 'Test Package',
            'file_pattern': 'MBB_*.csv',
            'source_file': self.source.id,
            'canvas_file': self.canvas.id,
            'input_format': 'csv',
            'output_format': 'xlsx',
            'output_prefix': 'mbb_output_',
            'batch_mode': 'instant',
            'batch_interval_minutes': 0,
        }
        data.update(overrides)
        return self.client.post('/api/transformation/packages/create/', data, format='json')

    # ── Create ──

    def test_create_package(self):
        resp = self._create_package()
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.data['name'], 'Test Package')
        self.assertEqual(resp.data['file_pattern'], 'MBB_*.csv')
        self.assertEqual(resp.data['status'], 'inactive')
        self.assertEqual(resp.data['mapping_status'], 'unmapped')
        self.assertEqual(resp.data['source_file_name'], 'bank.csv')
        self.assertEqual(resp.data['canvas_file_name'], 'output.xlsx')
        self.assertEqual(resp.data['created_by_name'], 'John Doe')
        self.assertEqual(resp.data['created_by_email'], 'pkguser@etl.local')

    def test_create_package_scheduled(self):
        resp = self._create_package(batch_mode='scheduled', batch_interval_minutes=120)
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.data['batch_mode'], 'scheduled')
        self.assertEqual(resp.data['batch_interval_minutes'], 120)

    def test_create_package_missing_name(self):
        resp = self._create_package(name='')
        self.assertEqual(resp.status_code, 400)

    # ── List ──

    def test_list_packages(self):
        self._create_package(name='Package A')
        self._create_package(name='Package B')
        resp = self.client.get('/api/transformation/packages/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.data), 2)

    def test_list_packages_filter_status(self):
        self._create_package(name='Package A')
        resp = self.client.get('/api/transformation/packages/?status=inactive')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.data), 1)
        resp2 = self.client.get('/api/transformation/packages/?status=active')
        self.assertEqual(len(resp2.data), 0)

    # ── Detail / Update / Delete ──

    def test_get_package_detail(self):
        create_resp = self._create_package()
        pkg_id = create_resp.data['id']
        resp = self.client.get(f'/api/transformation/packages/{pkg_id}/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['name'], 'Test Package')
        # Check source/canvas headers are included
        self.assertEqual(resp.data['source_headers'], ['Date', 'Description', 'Amount', 'Balance'])
        self.assertEqual(resp.data['canvas_headers'], ['TransDate', 'Desc', 'Value'])

    def test_update_package(self):
        create_resp = self._create_package()
        pkg_id = create_resp.data['id']
        resp = self.client.patch(
            f'/api/transformation/packages/{pkg_id}/',
            {'name': 'Updated Package', 'output_prefix': 'new_prefix_'},
            format='json',
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['name'], 'Updated Package')
        self.assertEqual(resp.data['output_prefix'], 'new_prefix_')

    def test_delete_package(self):
        create_resp = self._create_package()
        pkg_id = create_resp.data['id']
        resp = self.client.delete(f'/api/transformation/packages/{pkg_id}/')
        self.assertEqual(resp.status_code, 204)

    # ── Status Control ──

    def test_start_package(self):
        create_resp = self._create_package()
        pkg_id = create_resp.data['id']
        resp = self.client.post(f'/api/transformation/packages/{pkg_id}/start/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['status'], 'active')

    def test_pause_package(self):
        create_resp = self._create_package()
        pkg_id = create_resp.data['id']
        # Start first
        self.client.post(f'/api/transformation/packages/{pkg_id}/start/')
        resp = self.client.post(f'/api/transformation/packages/{pkg_id}/pause/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['status'], 'paused')

    def test_stop_package(self):
        create_resp = self._create_package()
        pkg_id = create_resp.data['id']
        self.client.post(f'/api/transformation/packages/{pkg_id}/start/')
        resp = self.client.post(f'/api/transformation/packages/{pkg_id}/stop/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['status'], 'inactive')

    def test_invalid_package_action(self):
        create_resp = self._create_package()
        pkg_id = create_resp.data['id']
        resp = self.client.post(f'/api/transformation/packages/{pkg_id}/restart/')
        self.assertEqual(resp.status_code, 400)

    def test_package_not_found(self):
        resp = self.client.post('/api/transformation/packages/99999/start/')
        self.assertEqual(resp.status_code, 404)


# ─── Field Mapping Tests ───────────────────────────────────────────────

@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class TestFieldMappingAPI(TestCase):
    def setUp(self):
        from oauth2_provider.models import Application, AccessToken
        from django.utils import timezone
        from apps.transformation.models import Package
        import datetime

        self.user = User.objects.create_user(
            username='mapper', email='mapper@etl.local', password='Test@12345',
        )
        self.app = Application.objects.create(
            name='test-app-map', client_type=Application.CLIENT_CONFIDENTIAL,
            authorization_grant_type=Application.GRANT_PASSWORD, user=self.user,
        )
        self.token = AccessToken.objects.create(
            user=self.user, application=self.app,
            token='test-token-map-101',
            expires=timezone.now() + datetime.timedelta(hours=1),
            scope='read write',
        )
        self.client = APIClient()
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {self.token.token}')

        self.source = UploadedFile.objects.create(
            file_type='source', original_filename='src.csv', file_format='csv',
            headers_json=['ColA', 'ColB', 'ColC', 'ColD'],
            field_count=4, uploaded_by=self.user,
        )
        self.canvas = UploadedFile.objects.create(
            file_type='canvas', original_filename='out.csv', file_format='csv',
            headers_json=['TargetX', 'TargetY', 'TargetZ'],
            field_count=3, uploaded_by=self.user,
        )
        self.package = Package.objects.create(
            name='Map Test Pkg', file_pattern='TEST_*.csv',
            source_file=self.source, canvas_file=self.canvas,
            input_format='csv', output_format='csv', output_prefix='test_',
            batch_mode='instant', created_by=self.user,
        )

    def test_save_mappings(self):
        mappings = [
            {'source_header': 'ColA', 'canvas_header': 'TargetX', 'order': 0},
            {'source_header': 'ColB', 'canvas_header': 'TargetY', 'order': 1},
            {'source_header': 'ColC', 'canvas_header': 'TargetZ', 'order': 2},
        ]
        resp = self.client.post(
            f'/api/transformation/packages/{self.package.id}/mappings/',
            mappings, format='json',
        )
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(len(resp.data), 3)
        # Verify mapping_status changed to mapped
        self.package.refresh_from_db()
        self.assertEqual(self.package.mapping_status, 'mapped')

    def test_list_mappings(self):
        from apps.transformation.models import FieldMapping
        FieldMapping.objects.create(
            package=self.package, source_header='ColA',
            canvas_header='TargetX', order=0,
        )
        resp = self.client.get(f'/api/transformation/packages/{self.package.id}/mappings/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.data), 1)
        self.assertEqual(resp.data[0]['source_header'], 'ColA')

    def test_replace_mappings(self):
        """Saving new mappings should replace existing ones."""
        from apps.transformation.models import FieldMapping
        FieldMapping.objects.create(
            package=self.package, source_header='ColA',
            canvas_header='TargetX', order=0,
        )
        # Replace with different mappings
        new_mappings = [
            {'source_header': 'ColD', 'canvas_header': 'TargetX', 'order': 0},
            {'source_header': 'ColC', 'canvas_header': 'TargetZ', 'order': 1},
        ]
        resp = self.client.post(
            f'/api/transformation/packages/{self.package.id}/mappings/',
            new_mappings, format='json',
        )
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(len(resp.data), 2)
        # Old mapping should be gone
        self.assertEqual(FieldMapping.objects.filter(package=self.package).count(), 2)

    def test_invalid_mapping_format(self):
        """Sending non-list should return 400."""
        resp = self.client.post(
            f'/api/transformation/packages/{self.package.id}/mappings/',
            {'source_header': 'ColA', 'canvas_header': 'TargetX'},
            format='json',
        )
        self.assertEqual(resp.status_code, 400)

    def test_mapping_not_found_package(self):
        resp = self.client.get('/api/transformation/packages/99999/mappings/')
        self.assertEqual(resp.status_code, 404)
