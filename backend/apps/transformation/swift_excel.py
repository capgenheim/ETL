"""
SWIFT message → Excel converter.
Generates formatted Excel workbooks from parsed SWIFT message JSON.
"""
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Style constants ───────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='1B2838')
HEADER_FONT = Font(name='Consolas', size=11, bold=True, color='FFFFFF')
ROW_FONT = Font(name='Consolas', size=10, color='000000')
ROW_FONT_TAG = Font(name='Consolas', size=10, color='1565C0', bold=True)
ROW_FONT_QUAL = Font(name='Consolas', size=10, color='0D47A1')
ROW_FONT_DESC = Font(name='Consolas', size=10, color='616161')
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='1B2838')
SUBTITLE_FONT = Font(name='Calibri', size=11, color='424242')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='BDBDBD'),
)
DATA_FILL_ODD = PatternFill('solid', fgColor='F5F5F5')
DATA_FILL_EVEN = PatternFill('solid', fgColor='FFFFFF')
SEQ_FILL = PatternFill('solid', fgColor='E3F2FD')
SEQ_FONT = Font(name='Consolas', size=10, color='1565C0', bold=True)


def generate_excel(parsed_data, output_dir, source_filename=''):
    """
    Generate an Excel file from parsed SWIFT message data.

    Args:
        parsed_data: dict from swift_parser.parse_swift_message()
        output_dir: directory for output file
        source_filename: original filename for naming

    Returns:
        str: full path to generated Excel file
    """
    mt = parsed_data.get('message_type', 'UNKNOWN')
    desc = parsed_data.get('message_type_description', '')
    sender = parsed_data.get('sender_bic', '')
    receiver = parsed_data.get('receiver_bic', '')
    reference = parsed_data.get('reference', '')
    fields = parsed_data.get('fields', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = mt

    # ── Sheet setup ──
    ws.sheet_properties.tabColor = 'FFC107'

    # ── Row 1: Title ──
    title = f'{mt} — {desc}' if desc else mt
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    # ── Row 2: Sender → Receiver ──
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Sender: {sender}  →  Receiver: {receiver}'
    ws['A2'].font = SUBTITLE_FONT

    # ── Row 3: Reference + filename ──
    ws.merge_cells('A3:F3')
    ws['A3'] = f'Reference: {reference}    |    Source: {source_filename}'
    ws['A3'].font = SUBTITLE_FONT

    # ── Row 4: blank ──
    ws.row_dimensions[4].height = 8

    # ── Row 5: Column headers ──
    headers = ['#', 'Tag', 'Value', 'Field Name', 'Description', 'Sequence']
    col_widths = [5, 8, 50, 25, 45, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[5].height = 22

    # ── Data rows ──
    row = 6
    seq_num = 0
    last_sequence = ''

    for field in fields:
        tag = field.get('tag', '')
        display = field.get('display', '')
        field_name = field.get('field_name', '')
        description = field.get('description', '')
        sequence = field.get('sequence', '')

        # Skip block delimiters from numbering but still show them
        is_block_delim = tag in ('16R', '16S')

        if is_block_delim:
            # Sequence row — visual separator
            if tag == '16R':
                ws.merge_cells(f'A{row}:F{row}')
                ws.cell(row=row, column=1, value=f'▸ {display}')
                ws.cell(row=row, column=1).font = SEQ_FONT
                ws.cell(row=row, column=1).fill = SEQ_FILL
                ws.row_dimensions[row].height = 20
                row += 1
            # Skip 16S rows entirely
            continue

        seq_num += 1
        fill = DATA_FILL_ODD if seq_num % 2 == 1 else DATA_FILL_EVEN

        ws.cell(row=row, column=1, value=seq_num).font = ROW_FONT_DESC
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

        tag_cell = ws.cell(row=row, column=2, value=f':{tag}:')
        tag_cell.font = ROW_FONT_TAG
        tag_cell.fill = fill

        val_cell = ws.cell(row=row, column=3, value=display)
        val_cell.font = ROW_FONT
        val_cell.fill = fill
        val_cell.alignment = Alignment(wrap_text=True)

        name_cell = ws.cell(row=row, column=4, value=field_name)
        name_cell.font = ROW_FONT
        name_cell.fill = fill

        desc_cell = ws.cell(row=row, column=5, value=description)
        desc_cell.font = ROW_FONT_DESC
        desc_cell.fill = fill

        seq_cell = ws.cell(row=row, column=6, value=sequence)
        seq_cell.font = ROW_FONT_DESC
        seq_cell.fill = fill

        # Bottom border for row
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER

        row += 1

    # ── Freeze panes & auto-filter ──
    ws.freeze_panes = 'A6'
    ws.auto_filter.ref = f'A5:F{row - 1}'

    # ── Generate filename ──
    ref_safe = reference.replace('/', '_').replace('\\', '_')[:30]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_name = f'{mt}_{ref_safe}_{timestamp}.xlsx'
    excel_name = re.sub(r'[^\w._-]', '_', excel_name)

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, excel_name)
    wb.save(filepath)

    return filepath, excel_name


# Need re for filename sanitization
import re
