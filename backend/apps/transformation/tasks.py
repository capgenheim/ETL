"""
Celery tasks for ETL file processing.
- file_sense_scan: Periodic task that scans trfm_inbound for new files matching active packages.
- process_inbound_file: Processes a single inbound file through its mapped package.
"""
import csv
import fnmatch
import io
import os
import shutil
from datetime import datetime

import openpyxl
import xlrd

from celery import shared_task
from django.conf import settings
from django.utils import timezone


@shared_task(name='transformation.file_sense_scan')
def file_sense_scan():
    """
    Scan trfm_inbound directory for files matching active package patterns.
    For each match, dispatch process_inbound_file task.
    """
    from .models import Package

    inbound_dir = settings.TRFM_INBOUND_DIR

    if not os.path.exists(inbound_dir):
        return {'scanned': 0, 'matched': 0}

    files = [f for f in os.listdir(inbound_dir) if os.path.isfile(os.path.join(inbound_dir, f))]
    if not files:
        return {'scanned': 0, 'matched': 0}

    # Get all active/running packages with mappings
    active_packages = Package.objects.filter(
        status__in=[Package.Status.ACTIVE, Package.Status.RUNNING],
        mapping_status=Package.MappingStatus.MAPPED,
    ).select_related('source_file', 'canvas_file')

    matched = 0
    for filename in files:
        for package in active_packages:
            if fnmatch.fnmatch(filename.lower(), package.file_pattern.lower()):
                # Dispatch processing task
                filepath = os.path.join(inbound_dir, filename)
                process_inbound_file.delay(package.id, filepath, filename, 'instant')
                matched += 1
                break  # One file matches one package

    return {'scanned': len(files), 'matched': matched}


@shared_task(name='transformation.process_inbound_file', bind=True, max_retries=3)
def process_inbound_file(self, package_id, filepath, original_filename, run_type='instant'):
    """
    Process a single inbound file:
    1. Read the source file
    2. Apply field mappings — unmapped canvas columns output blank
    3. Write transformed output to trfm_outbound
    4. Save inbound file to PostgreSQL (InboundFileLog)
    5. Delete inbound file from disk
    """
    from .models import Package, InboundFileLog

    try:
        package = Package.objects.select_related(
            'source_file', 'canvas_file'
        ).get(pk=package_id)
    except Package.DoesNotExist:
        return {'error': f'Package {package_id} not found'}

    if not os.path.exists(filepath):
        return {'error': f'File not found: {filepath}'}

    # Mark package as running
    package.status = Package.Status.RUNNING
    package.save(update_fields=['status', 'updated_at'])

    try:
        # Get ALL canvas headers from the canvas file
        all_canvas_headers = package.canvas_file.headers_json or []

        # Get field mappings as dict keyed by canvas_header
        mappings_qs = list(
            package.field_mappings.all()
            .order_by('order')
            .values('source_header', 'canvas_header', 'mapping_type',
                    'has_condition', 'condition_json', 'constant_value')
        )
        mapping_by_canvas = {m['canvas_header']: m for m in mappings_qs}

        # Read source data
        source_data = _read_file(filepath, original_filename)

        if not source_data:
            return {'error': 'No data in source file'}

        # Build full mappings list: mapped fields use their config, unmapped → blank
        full_mappings = []
        for ch in all_canvas_headers:
            if ch in mapping_by_canvas:
                full_mappings.append(mapping_by_canvas[ch])
            else:
                # Unmapped canvas column → blank for every row
                full_mappings.append({
                    'canvas_header': ch,
                    'source_header': '',
                    'mapping_type': 'constant',
                    'has_condition': False,
                    'condition_json': None,
                    'constant_value': '',
                })

        # Transform: remap columns (includes blanks for unmapped)
        transformed = _transform_data(source_data, full_mappings)

        # Generate output filename: <prefix><ddmmyyyyhhmmss>.<format>
        now = timezone.localtime()
        timestamp = now.strftime('%d%m%Y%H%M%S')
        output_filename = f'{package.output_prefix}{timestamp}.{package.output_format}'
        output_path = os.path.join(settings.TRFM_OUTBOUND_DIR, output_filename)

        # Write output
        _write_file(output_path, transformed, package.output_format)

        # Read inbound file binary content before deleting
        with open(filepath, 'rb') as f:
            file_bytes = f.read()

        file_ext = os.path.splitext(original_filename)[1].lower().lstrip('.')
        rows_processed = len(transformed) - 1  # Exclude header row

        # Save inbound file to PostgreSQL
        log_entry = InboundFileLog.objects.create(
            package=package,
            original_filename=original_filename,
            file_content=file_bytes,
            file_size=len(file_bytes),
            run_type=run_type,
            file_format=file_ext or 'csv',
            output_filename=output_filename,
            rows_processed=rows_processed,
            status=InboundFileLog.Status.SUCCESS,
        )

        # Auto-create tags for this file log
        from .models import FileTag
        auto_tags = [
            ((file_ext or 'csv').upper(), '#2196F3'),    # File format tag (blue)
            (package.name, '#FF9800'),                    # Package name tag (orange)
            (run_type.replace('_', ' ').title(), '#4CAF50'),  # Run type tag (green)
            ('Success', '#00C853'),                       # Status tag
        ]
        for tag_name, tag_color in auto_tags:
            tag, _ = FileTag.objects.get_or_create(
                name=tag_name,
                defaults={'color': tag_color, 'auto_created': True}
            )
            log_entry.tags.add(tag)

        # Delete inbound file from disk
        os.remove(filepath)

        # Restore package status to active
        package.status = Package.Status.ACTIVE
        package.save(update_fields=['status', 'updated_at'])

        return {
            'success': True,
            'output_file': output_filename,
            'rows_processed': rows_processed,
            'stored_in_db': True,
        }

    except Exception as exc:
        # Log failure to DB if possible
        try:
            from .models import InboundFileLog
            InboundFileLog.objects.create(
                package=package,
                original_filename=original_filename,
                file_content=b'',
                file_size=0,
                status=InboundFileLog.Status.FAILED,
                run_type=run_type,
                error_message=str(exc),
            )
        except Exception:
            pass

        # Restore status on failure
        package.status = Package.Status.ACTIVE
        package.save(update_fields=['status', 'updated_at'])
        raise self.retry(exc=exc, countdown=30)


def _read_file(filepath, filename):
    """Read file into list of dicts based on file extension."""
    ext = os.path.splitext(filename)[1].lower().lstrip('.')

    if ext == 'csv':
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            return list(reader)

    elif ext == 'xlsx':
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(h) if h else '' for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]

    elif ext == 'xls':
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        if ws.nrows < 2:
            return []
        headers = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
        return [
            dict(zip(headers, [ws.cell_value(r, c) for c in range(ws.ncols)]))
            for r in range(1, ws.nrows)
        ]

    return []


def _transform_data(source_data, mappings):
    """Apply field mappings to transform source data to canvas schema.
    Supports direct mapping, conditional IF/ELSE, and constant values.
    mappings is a list of dicts with keys: source_header, canvas_header,
    mapping_type, has_condition, condition_json, constant_value.
    """
    canvas_headers = [m['canvas_header'] for m in mappings]
    transformed = [canvas_headers]  # Header row

    for row in source_data:
        new_row = []
        for m in mappings:
            mtype = m.get('mapping_type', 'direct')

            # ── Constant: fixed value for every row ──
            if mtype == 'constant':
                new_row.append(m.get('constant_value', ''))

            # ── Conditional: evaluate IF / IF-ELSE ──
            elif mtype == 'condition' and m.get('has_condition') and m.get('condition_json'):
                cond = m['condition_json']
                try:
                    result = _evaluate_condition(row, cond)
                    new_row.append(result if result is not None else '')
                except Exception:
                    new_row.append('')

            # ── Direct: simple column copy ──
            else:
                value = row.get(m.get('source_header', ''), '')
                new_row.append(value if value is not None else '')

        transformed.append(new_row)

    return transformed


def _evaluate_condition(row, cond):
    """Evaluate a single IF or IF-ELSE condition against a data row.

    cond schema:
    {
        "condition_type": "if_else" | "if_only" | "fixed_value",
        "source_field": "Debit (DR)",   # field to test (empty for fixed_value)
        "operator": ">",                 # comparison operator (empty for fixed_value)
        "compare_value": "0",            # value to compare against
        "then_source": "Debit (DR)",     # field to use if true (optional)
        "then_value": "DEBIT",           # literal value if true / fixed value
        "else_source": "Credit (CR)",    # field to use if false (optional)
        "else_value": "",                # literal value if false (optional)
    }
    """
    condition_type = cond.get('condition_type', 'if_else')

    # Fixed Value: return then_value directly, no condition evaluation
    if condition_type == 'fixed_value':
        return cond.get('then_value', '')

    source_field = cond.get('source_field', '')
    operator = cond.get('operator', '==')
    compare_value = cond.get('compare_value', '')

    cell_raw = row.get(source_field, '')
    if cell_raw is None:
        cell_raw = ''

    # Evaluate condition
    condition_met = _compare(cell_raw, operator, compare_value)

    if condition_met:
        # THEN branch: use then_source field or then_value literal
        if 'then_source' in cond and cond['then_source']:
            val = row.get(cond['then_source'], '')
            return val if val is not None else ''
        return cond.get('then_value', '')
    else:
        if condition_type == 'if_only':
            return ''  # No else branch — leave blank
        # ELSE branch: use else_source field or else_value literal
        if 'else_source' in cond and cond['else_source']:
            val = row.get(cond['else_source'], '')
            return val if val is not None else ''
        return cond.get('else_value', '')


def _compare(cell_raw, operator, compare_value):
    """Compare cell_raw against compare_value using the given operator."""
    # Try numeric comparison first
    try:
        num_cell = float(str(cell_raw))
        num_comp = float(str(compare_value))
        if operator == '>':
            return num_cell > num_comp
        elif operator == '<':
            return num_cell < num_comp
        elif operator == '>=':
            return num_cell >= num_comp
        elif operator == '<=':
            return num_cell <= num_comp
        elif operator == '==':
            return num_cell == num_comp
        elif operator == '!=':
            return num_cell != num_comp
    except (ValueError, TypeError):
        pass

    # String-based comparison
    str_cell = str(cell_raw).strip()
    str_comp = str(compare_value).strip()

    if operator == '==':
        return str_cell == str_comp
    elif operator == '!=':
        return str_cell != str_comp
    elif operator == 'contains':
        return str_comp.lower() in str_cell.lower()
    elif operator == 'not_empty':
        return bool(str_cell)
    elif operator == 'is_empty':
        return not bool(str_cell)

    return False


def _write_file(output_path, data, fmt):
    """Write transformed data to output file in the specified format."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if fmt == 'csv':
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerows(data)

    elif fmt == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transformed'
        for row in data:
            ws.append(row)
        wb.save(output_path)

    elif fmt == 'xls':
        import xlwt
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Transformed')
        for r, row in enumerate(data):
            for c, val in enumerate(row):
                ws.write(r, c, val)
        wb.save(output_path)


# ═══════════════════════════════════════════════════════════════════════
# SWIFT Inbound Pipeline Tasks (MT + MX)
# ═══════════════════════════════════════════════════════════════════════


def _get_or_create_tag(name, color='#40c4ff'):
    """Get or create a FileTag with auto_created=True."""
    from .models import FileTag
    tag, _ = FileTag.objects.get_or_create(
        name=name,
        defaults={'color': color, 'auto_created': True},
    )
    return tag


def _register_swift_file_log(original_filename, raw_bytes, output_filename='',
                              status_val='success', error_message='',
                              msg_type='', is_mx=False):
    """
    Create an InboundFileLog entry for a SWIFT file with auto tags.
    This makes SWIFT files visible in File Manager (Source Files / Unprocessed Files).
    """
    from .models import InboundFileLog

    file_ext = os.path.splitext(original_filename)[1].lower().lstrip('.')

    log = InboundFileLog.objects.create(
        package=None,  # SWIFT files don't belong to an ETL package
        original_filename=original_filename,
        file_content=raw_bytes,
        file_size=len(raw_bytes),
        file_format=file_ext or 'txt',
        output_filename=output_filename,
        rows_processed=0,
        status=status_val,
        error_message=error_message,
        run_type='swift',
    )

    # Auto-create and assign tags: #SWIFT, #MT or #MX, and the message type
    tags_to_add = [
        _get_or_create_tag('#SWIFT', '#00BCD4'),
    ]
    if is_mx:
        tags_to_add.append(_get_or_create_tag('#MX', '#4FC3F7'))
    else:
        tags_to_add.append(_get_or_create_tag('#MT', '#FF9800'))

    if msg_type:
        tags_to_add.append(_get_or_create_tag(f'#{msg_type}', '#7C4DFF'))

    log.tags.add(*tags_to_add)
    return log


@shared_task(name='transformation.swift_file_sense_scan')
def swift_file_sense_scan():
    """
    Scan sft_inbound directory for SWIFT MT (.txt/.fin) and MX (.xml) files.
    For each file found, dispatch process_swift_file task.
    """
    from .swift_parser import is_swift_message
    from .mx_parser import is_mx_message

    inbound_dir = settings.SFT_INBOUND_DIR

    if not os.path.exists(inbound_dir):
        return {'scanned': 0, 'dispatched': 0}

    files = [
        f for f in os.listdir(inbound_dir)
        if os.path.isfile(os.path.join(inbound_dir, f))
        and f.lower().endswith(('.txt', '.fin', '.swift', '.mt', '.xml'))
    ]

    if not files:
        return {'scanned': 0, 'dispatched': 0}

    dispatched = 0
    for filename in files:
        filepath = os.path.join(inbound_dir, filename)
        try:
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
            if is_swift_message(content) or is_mx_message(content):
                process_swift_file.delay(filepath, filename)
                dispatched += 1
        except Exception:
            pass

    return {'scanned': len(files), 'dispatched': dispatched}


@shared_task(name='transformation.process_swift_file', bind=True, max_retries=2)
def process_swift_file(self, filepath, original_filename):
    """
    Process a single SWIFT inbound file (MT or MX):
    1. Read file content
    2. Detect format (MT FIN vs MX XML)
    3. Parse using appropriate parser
    4. Save SwiftMessage record to DB
    5. Generate Excel to sft_outbound
    6. Register in InboundFileLog with auto tags
    7. Remove source file from sft_inbound
    """
    from .models import SwiftMessage
    from .swift_parser import parse_swift_message, is_swift_message
    from .mx_parser import parse_mx_message, is_mx_message
    from .swift_excel import generate_excel

    outbound_dir = settings.SFT_OUTBOUND_DIR

    try:
        # 1. Read file
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            raw_content = f.read()

        raw_bytes = raw_content.encode('utf-8')

        # 2. Detect format and split/parse
        is_mx = is_mx_message(raw_content)
        if is_mx:
            # MX (ISO 20022 XML)
            messages_raw = _split_mx_messages(raw_content)
            parser = parse_mx_message
        else:
            # MT (FIN)
            messages_raw = _split_swift_messages(raw_content)
            parser = parse_swift_message

        results = []
        for msg_text in messages_raw:
            # 3. Parse
            parsed = parser(msg_text)

            # 4. Generate Excel
            excel_path, excel_name = generate_excel(parsed, outbound_dir, original_filename)

            # 5. Save to DB
            msg = SwiftMessage.objects.create(
                message_type=parsed['message_type'],
                message_type_description=parsed['message_type_description'],
                sender_bic=parsed['sender_bic'],
                receiver_bic=parsed['receiver_bic'],
                reference=parsed['reference'],
                raw_content=msg_text,
                parsed_json=parsed,
                original_filename=original_filename,
                excel_filename=excel_name,
                source_file_content=raw_bytes,
                status='processed',
                processed_at=timezone.now(),
            )

            # 6. Register in InboundFileLog for File Manager
            _register_swift_file_log(
                original_filename=original_filename,
                raw_bytes=raw_bytes,
                output_filename=excel_name,
                status_val='success',
                msg_type=parsed['message_type'],
                is_mx=is_mx,
            )

            results.append({
                'id': msg.id,
                'type': parsed['message_type'],
                'ref': parsed['reference'],
                'excel': excel_name,
            })

        # 7. Remove source file
        if os.path.exists(filepath):
            os.remove(filepath)

        return {
            'status': 'success',
            'filename': original_filename,
            'messages_processed': len(results),
            'details': results,
        }

    except Exception as exc:
        # Save failed record
        try:
            from .models import SwiftMessage
            raw = open(filepath, 'r', errors='replace').read() if os.path.exists(filepath) else ''
            SwiftMessage.objects.create(
                message_type='UNKNOWN',
                raw_content=raw,
                original_filename=original_filename,
                status='failed',
                error_message=str(exc),
                processed_at=timezone.now(),
            )
            # Register failed file in InboundFileLog (shows in Unprocessed Files)
            _register_swift_file_log(
                original_filename=original_filename,
                raw_bytes=raw.encode('utf-8'),
                status_val='failed',
                error_message=str(exc),
            )
        except Exception:
            pass

        raise self.retry(exc=exc, countdown=10)


def _split_swift_messages(content):
    """
    Split file content into individual SWIFT MT messages.
    Each message starts with {1:F and ends with -}
    """
    import re
    pattern = r'(\{1:F.*?-\})'
    matches = re.findall(pattern, content, re.DOTALL)

    if matches:
        return matches

    return [content.strip()] if content.strip() else []


def _split_mx_messages(content):
    """
    Split file content into individual MX (DataPDU XML) messages.
    Each message is a <DataPDU>...</DataPDU> block or a full XML document.
    """
    import re

    # Try splitting on <DataPDU> blocks
    pattern = r'(<\?xml[^?]*\?>.*?</DataPDU>)'
    matches = re.findall(pattern, content, re.DOTALL)

    if matches:
        return matches

    # Try without the XML declaration
    pattern2 = r'(<DataPDU[^>]*>.*?</DataPDU>)'
    matches2 = re.findall(pattern2, content, re.DOTALL)

    if matches2:
        return matches2

    return [content.strip()] if content.strip() else []


