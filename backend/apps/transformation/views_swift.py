"""
SWIFT Parameter CRUD API views.
"""
import json
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import SwiftParameter


class SwiftParameterListCreateView(APIView):
    """
    GET  /api/transformation/swift-params/       — List with filters
    POST /api/transformation/swift-params/       — Create new parameter
    """

    def get(self, request):
        qs = SwiftParameter.objects.all()

        # Filters
        category = request.query_params.get('category')
        if category:
            qs = qs.filter(category=category)

        message_type = request.query_params.get('message_type')
        if message_type:
            qs = qs.filter(message_type__icontains=message_type)

        status_filter = request.query_params.get('status')
        if status_filter in ('active', 'inactive'):
            qs = qs.filter(status=status_filter)

        search = request.query_params.get('search', '').strip()
        if search:
            from django.db.models import Q
            qs = qs.filter(
                Q(field_tag__icontains=search) |
                Q(field_name__icontains=search) |
                Q(message_type__icontains=search) |
                Q(description__icontains=search)
            )

        data_type = request.query_params.get('data_type')
        if data_type:
            qs = qs.filter(data_type=data_type)

        data = [
            {
                'id': p.id,
                'category': p.category,
                'message_type': p.message_type,
                'field_tag': p.field_tag,
                'field_name': p.field_name,
                'description': p.description,
                'data_type': p.data_type,
                'is_mandatory': p.is_mandatory,
                'max_length': p.max_length,
                'format_pattern': p.format_pattern,
                'sample_value': p.sample_value,
                'options_json': p.options_json,
                'status': p.status,
                'created_at': p.created_at.isoformat(),
                'updated_at': p.updated_at.isoformat(),
            }
            for p in qs[:500]
        ]
        return Response(data)

    def post(self, request):
        fields = [
            'category', 'message_type', 'field_tag', 'field_name',
            'description', 'data_type', 'is_mandatory', 'max_length',
            'format_pattern', 'sample_value', 'options_json', 'status',
        ]
        kwargs = {}
        for f in fields:
            if f in request.data:
                kwargs[f] = request.data[f]

        # Defaults
        kwargs.setdefault('status', 'active')
        kwargs.setdefault('data_type', 'TEXT')
        kwargs['created_by'] = request.user

        try:
            param = SwiftParameter.objects.create(**kwargs)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'id': param.id,
            'message': f'Created: {param.message_type} {param.field_tag}',
        }, status=status.HTTP_201_CREATED)


class SwiftParameterDetailView(APIView):
    """
    GET    /api/transformation/swift-params/<id>/  — Detail
    PUT    /api/transformation/swift-params/<id>/  — Update
    DELETE /api/transformation/swift-params/<id>/  — Delete
    """

    def get_object(self, pk):
        try:
            return SwiftParameter.objects.get(pk=pk)
        except SwiftParameter.DoesNotExist:
            return None

    def get(self, request, pk):
        param = self.get_object(pk)
        if not param:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        return Response({
            'id': param.id,
            'category': param.category,
            'message_type': param.message_type,
            'field_tag': param.field_tag,
            'field_name': param.field_name,
            'description': param.description,
            'data_type': param.data_type,
            'is_mandatory': param.is_mandatory,
            'max_length': param.max_length,
            'format_pattern': param.format_pattern,
            'sample_value': param.sample_value,
            'options_json': param.options_json,
            'status': param.status,
            'created_at': param.created_at.isoformat(),
            'updated_at': param.updated_at.isoformat(),
        })

    def put(self, request, pk):
        param = self.get_object(pk)
        if not param:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        updatable = [
            'category', 'message_type', 'field_tag', 'field_name',
            'description', 'data_type', 'is_mandatory', 'max_length',
            'format_pattern', 'sample_value', 'options_json', 'status',
        ]
        for field in updatable:
            if field in request.data:
                setattr(param, field, request.data[field])

        try:
            param.save()
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'id': param.id, 'message': 'Updated successfully'})

    def delete(self, request, pk):
        param = self.get_object(pk)
        if not param:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        name = f'{param.message_type} {param.field_tag}'
        param.delete()
        return Response({'message': f'Deleted: {name}'})


class SwiftParameterExportView(APIView):
    """
    GET /api/transformation/swift-params/export/ — Download all parameters as CSV
    """

    def get(self, request):
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="swift_parameters.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'category', 'message_type', 'field_tag', 'field_name',
            'description', 'data_type', 'is_mandatory', 'max_length',
            'format_pattern', 'sample_value', 'status',
        ])

        for p in SwiftParameter.objects.all():
            writer.writerow([
                p.category, p.message_type, p.field_tag, p.field_name,
                p.description, p.data_type, p.is_mandatory, p.max_length or '',
                p.format_pattern, p.sample_value, p.status,
            ])

        return response


class SwiftParameterImportView(APIView):
    """
    POST /api/transformation/swift-params/import/ — Upload CSV/Excel to import parameters
    """

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        fname = file.name.lower()
        rows = []

        try:
            if fname.endswith('.csv'):
                import csv
                import io
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif fname.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                headers = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v) if v is not None else '' for v in row])))
            else:
                return Response(
                    {'error': 'Unsupported file format. Use .csv or .xlsx'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response({'error': f'File parse error: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        required = ['category', 'message_type', 'field_tag', 'field_name']
        if not rows:
            return Response({'error': 'File is empty'}, status=status.HTTP_400_BAD_REQUEST)

        missing = [f for f in required if f not in rows[0]]
        if missing:
            return Response(
                {'error': f'Missing columns: {", ".join(missing)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        updated = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            try:
                is_mandatory = str(row.get('is_mandatory', 'False')).strip()
                is_mandatory = is_mandatory.lower() in ('true', '1', 'yes')

                max_len = row.get('max_length', '').strip()
                max_len = int(max_len) if max_len and max_len.isdigit() else None

                obj, was_created = SwiftParameter.objects.update_or_create(
                    category=row['category'].strip(),
                    message_type=row['message_type'].strip(),
                    field_tag=row['field_tag'].strip(),
                    defaults={
                        'field_name': row.get('field_name', '').strip(),
                        'description': row.get('description', '').strip(),
                        'data_type': row.get('data_type', 'TEXT').strip() or 'TEXT',
                        'is_mandatory': is_mandatory,
                        'max_length': max_len,
                        'format_pattern': row.get('format_pattern', '').strip(),
                        'sample_value': row.get('sample_value', '').strip(),
                        'status': row.get('status', 'active').strip() or 'active',
                        'created_by': request.user,
                    }
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append(f'Row {i}: {e}')

        result = {
            'created': created,
            'updated': updated,
            'total': SwiftParameter.objects.count(),
            'message': f'Import complete: {created} created, {updated} updated',
        }
        if errors:
            result['errors'] = errors[:20]

        return Response(result, status=status.HTTP_201_CREATED)

